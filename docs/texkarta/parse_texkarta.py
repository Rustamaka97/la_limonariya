#!/usr/bin/env python3
"""Parse Limonariya tech-card xlsx files into a clean JSON seed.

Usage: python3 parse_texkarta.py  (run from docs/texkarta/)
Output: texkarta.json
"""
import json
import re
import openpyxl

DIR = '.'


def norm_num(s):
    return float(s.replace(',', '.'))


def parse_qty(raw):
    """Return dict: qty_g (grams) OR qty+unit for pieces, plus marinade note if embedded."""
    if raw is None:
        return {'raw': None, 'qty_g': None, 'unit': None}
    raw = str(raw).strip()
    out = {'raw': raw, 'qty_g': None, 'unit': None, 'marinade': None}
    if raw in ('—', '-', ''):
        out['unit'] = 'po_vkusu'
        return out

    main = raw
    if '/' in raw:  # marinade growth note: "1 кг / 30% купаяди ..."
        main, note = raw.split('/', 1)
        note = note.strip()
        m = re.search(r'(\d+(?:[.,]\d+)?)\s*%', note)
        if m:
            out['marinade'] = {'gain_pct': norm_num(m.group(1)), 'note': note}
        else:
            m = re.search(r'(\d+(?:[.,]\d+)?)\s*гр', note)
            if m:
                out['marinade'] = {'gain_g_per_base': norm_num(m.group(1)), 'note': note}
        main = main.strip()

    m = re.fullmatch(r'(\d+(?:[.,]\d+)?)\s*кг\.?', main.strip())
    if m:
        out['qty_g'] = round(norm_num(m.group(1)) * 1000, 1)
        out['unit'] = 'g'
        return out
    m = re.fullmatch(r'(\d+(?:[.,]\d+)?)\s*гр?\.?', main.strip())
    if m:
        out['qty_g'] = norm_num(m.group(1))
        out['unit'] = 'g'
        return out
    m = re.fullmatch(r'(\d+(?:[.,]\d+)?)\s*(штук|дона|сих)\.?', main.strip())
    if m:
        out['qty'] = norm_num(m.group(1))
        out['unit'] = 'piece'
        return out
    out['unit'] = 'unparsed'
    return out


YIELD_RE = re.compile(r'\(\s*(\d+)\s*(порция|штук|дона|сих)\s*\)')

STOCK_HINTS = {
    'думба': 'Думба', 'фарш': 'Фарш', 'бон филе': 'Бон-филе',
    'илик суяк': 'Илик суяк', 'суяк': 'Илик суяк?', 'чарви': 'Чарви-г',
    'корин чарви': 'Чарви-г?',
}


def stock_hint(ing_name, category):
    n = ing_name.strip().lower()
    if n in STOCK_HINTS:
        return STOCK_HINTS[n]
    if 'гўшт' in n or n == 'гушт':
        return {'мол': 'мол лаҳм (обвалка)', 'куй': 'қўй лаҳм (обвалка)',
                'товуқ': 'товуқ гўшт'}.get(category)
    return None


def parse_dishes():
    wb = openpyxl.load_workbook(f'{DIR}/техкарта limonariya.xlsx', data_only=True)
    ws = wb['Рецептлар']
    dishes, cur = [], None
    for row in ws.iter_rows(min_row=4, values_only=True):
        num, name, cat, ing, qty = (row + (None,) * 5)[:5]
        if not any(v not in (None, '') for v in (num, name, ing, qty)):
            continue
        if name:  # new dish
            clean = ' '.join(str(name).split())
            ym = YIELD_RE.search(clean)
            cur = {
                'num': int(num) if num else None,
                'name': YIELD_RE.sub('', clean).strip().rstrip(','),
                'type': 'hot',
                'category': str(cat).strip() if cat and str(cat).strip() not in ('—', '-') else None,
                'yield': {'qty': int(ym.group(1)), 'unit': ym.group(2)} if ym else None,
                'marinade': None,
                'ingredients': [],
            }
            dishes.append(cur)
        if cur is None or not ing:
            continue
        q = parse_qty(qty)
        if q.pop('marinade', None) is not None:
            cur['marinade'] = parse_qty(qty)['marinade']
        item = {'name': ' '.join(str(ing).split()), **q}
        hint = stock_hint(item['name'], cur['category'])
        if hint:
            item['stock_hint'] = hint
        cur['ingredients'].append(item)
    return dishes


def parse_salads():
    wb = openpyxl.load_workbook(f'{DIR}/Салатлар_тех_карта.xlsx', data_only=True)
    ws = wb['Салатлар тех карта']
    salads, cur = [], None
    for row in ws.iter_rows(min_row=2, values_only=True):
        a, b, c = (row + (None,) * 3)[:3]
        a_s = ' '.join(str(a).split()) if a else ''
        if not a_s:
            continue
        if c and 'Чиқиш' in str(c):  # salad header
            m = re.search(r'(\d+)', str(c))
            cur = {'name': a_s, 'type': 'salad',
                   'yield_g_stated': int(m.group(1)) if m else None,
                   'sum_g_stated': None, 'ingredients': []}
            salads.append(cur)
            continue
        if cur is None or a_s == 'Масаллиқ':
            continue
        if a_s.startswith('ЖАМИ'):
            cur['sum_g_stated'] = b if isinstance(b, (int, float)) else None
            continue
        qty_g = float(b) if isinstance(b, (int, float)) else None
        item = {'name': a_s, 'raw': str(b) if b is not None else None,
                'qty_g': qty_g, 'unit': 'g' if qty_g is not None else None}
        if 'мол гўшт' in a_s:
            item['stock_hint'] = 'мол лаҳм (обвалка)'
        elif 'товуқ' in a_s:
            item['stock_hint'] = 'товуқ гўшт'
        if c:
            item['note'] = ' '.join(str(c).split())
        cur['ingredients'].append(item)
    return salads


CHALOP = {
    'name': 'Чалоп Айрон', 'type': 'salad', 'yield_g_stated': None,
    'sum_g_stated': 302, 'source': 'Sherxon, Telegram 12.06.2026 (Рустам ака берди)',
    'ingredients': [
        {'name': 'чакка', 'raw': '200 гр', 'qty_g': 200, 'unit': 'g'},
        {'name': 'помидор', 'raw': '50 гр', 'qty_g': 50, 'unit': 'g'},
        {'name': 'бодринг', 'raw': '50 гр', 'qty_g': 50, 'unit': 'g'},
        {'name': 'туз', 'raw': '2 гр', 'qty_g': 2, 'unit': 'g'},
    ],
}

ALIASES = {  # canonical -> spellings seen in the files
    'бодринг': ['бодиринг'],
    'болгар қалампир': ['светофор перец', 'светафор перец', 'светофор қалампир', 'болгар қалампир (2 хил)'],
    'корейски туз': ['корейча туз'],
    'кашнич': ['кенза', 'кензо'],
    'майонез': ['майнез'],
    'пармезан сыр': ['пармизан сир'],
    'чакка': ['чакки'],
    'мурч': [],
    'масло': [],
}


def main():
    dishes = parse_dishes()
    salads = parse_salads() + [CHALOP]

    for s in salads:  # integrity check
        calc = sum(i['qty_g'] for i in s['ingredients'] if i.get('qty_g'))
        s['sum_g_calc'] = round(calc, 1)
        flags = []
        if s['sum_g_stated'] is not None and abs(calc - s['sum_g_stated']) > 0.5:
            flags.append(f"ЖАМИ {s['sum_g_stated']} ≠ ҳисобланган {calc}")
        if s['yield_g_stated'] is not None and abs(calc - s['yield_g_stated']) > 5:
            flags.append(f"чиқиш {s['yield_g_stated']} vs масаллиқлар {calc}")
        if flags:
            s['flags'] = flags

    ing_names = {}
    for d in dishes + salads:
        for i in d['ingredients']:
            ing_names.setdefault(i['name'].lower(), 0)
            ing_names[i['name'].lower()] += 1

    out = {
        'source': {
            'files': ['техкарта limonariya.xlsx', 'Салатлар_тех_карта.xlsx'],
            'received': '2026-06-12', 'from': 'Рустам ака',
            'note': 'Чалоп Айрон — қўлда қўшилган (файлда йўқ эди)',
        },
        'dishes': dishes,
        'salads': salads,
        'ingredient_usage': dict(sorted(ing_names.items(), key=lambda kv: -kv[1])),
        'aliases': ALIASES,
    }
    with open(f'{DIR}/texkarta.json', 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"dishes: {len(dishes)}, salads: {len(salads)}, "
          f"unique ingredients: {len(ing_names)}")
    for s in salads:
        if s.get('flags'):
            print(f"FLAG {s['name']}: {s['flags']}")


if __name__ == '__main__':
    main()
